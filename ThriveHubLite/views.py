from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Avg
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from calls.models import CallSession, Caller, SiteConfig, Holiday, ShiftType, ShiftAssignment, ShiftPreference, TimeOffRequest, MonthlyTeleconsult
from calls.forms import SiteConfigForm, UserProfileForm, HolidayForm, ShiftTypeForm, AssignTeleconsultForm
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from calls.utils import generate_calendar_matrix
from collections import defaultdict
from django.db.models.functions import ExtractYear
from django.urls import reverse
import calendar
import csv
import os
import pandas as pd
from django.views.decorators.http import require_POST
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from django.core.serializers.json import DjangoJSONEncoder
from calendar import monthrange, month_name
from collections import defaultdict


GENDER_LABELS = {
    'male': 'Male',
    'female': 'Female',
    'lgbtq': 'LGBTQ++',
    'na': 'N/A',
}
GENDER_COLORS = {
    'Male': '#4C8CAF',
    'Female': '#F48FB1',
    'LGBTQ++': '#F7A6B8',
    'N/A': '#CBDDD5'
}
month_names = list(enumerate(calendar.month_name))[1:]


def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + timedelta(n)


@login_required
def dashboard(request):
    # Filter: Year & Month from query params or fallback to current
    year = int(request.GET.get('year', datetime.now().year))
    month = request.GET.get('month')
    selected_month = int(month) if month else None
    selected_year = year

    current_start = datetime(year, selected_month or 1, 1)
    prev_start = current_start - relativedelta(months=1)

    # Logs this period
    logs = CallSession.objects.filter(date__year=year)
    if month:
        logs = logs.filter(date__month=month)

    # Logs previous period
    prev_logs = CallSession.objects.filter(
        date__year=prev_start.year,
        date__month=prev_start.month
    )

    # Total Calls
    total_calls = logs.count()
    prev_total_calls = prev_logs.count()
    delta_total = total_calls - prev_total_calls
    trend_total = "up" if delta_total > 0 else "down" if delta_total < 0 else "flat"

    # Most Common Age & Age Cluster
    common_age_data = logs.values('caller__age').annotate(count=Count('id')).order_by('-count').first()
    common_age = common_age_data['caller__age'] if common_age_data else None
    age_range = (common_age - 2, common_age + 2) if common_age else (None, None)

    # High Risk %
    high_risk = logs.filter(risk_level="high").count()
    high_risk_percent = round((high_risk / total_calls * 100), 1) if total_calls else 0

    # Gender distribution
    gender_data = logs.values('caller__gender').annotate(count=Count('id')).order_by('caller__gender')
    gender_labels = [GENDER_LABELS.get(g['caller__gender'], 'Other') for g in gender_data]
    gender_counts = [g['count'] for g in gender_data]
    gender_colors = [GENDER_COLORS.get(label, '#999999') for label in gender_labels]

    # Top 10 Reasons
    reasons_data = logs.values('reasons_for_calling__label') \
        .annotate(count=Count('id')).order_by('-count')[:10]
    top_reasons = []
    max_count = reasons_data[0]['count'] if reasons_data else 1
    base_colors = ['#4CAF8B', '#81C9B3', '#2F6657', '#CBDDD5', '#94BFB1']
    for i, item in enumerate(reasons_data):
        top_reasons.append({
            'label': item['reasons_for_calling__label'] or "Unspecified",
            'count': item['count'],
            'percent': int(item['count'] / max_count * 100),
            'color': base_colors[i % len(base_colors)],
        })

    # Monthly Shift Breakdown
    monthly_raw = logs.values('date__month', 'shift').annotate(count=Count('id'))
    monthly_shifts = defaultdict(lambda: {'6-2': 0, '2-10': 0, '10-6': 0, '8-4': 0})
    for item in monthly_raw:
        monthly_shifts[item['date__month']][item['shift']] = item['count']
    shift_chart_data = {
        'labels': [calendar.month_abbr[i] for i in range(1, 13)],
        '6-2': [monthly_shifts[i]['6-2'] for i in range(1, 13)],
        '2-10': [monthly_shifts[i]['2-10'] for i in range(1, 13)],
        '10-6': [monthly_shifts[i]['10-6'] for i in range(1, 13)],
        '8-4': [monthly_shifts[i]['8-4'] for i in range(1, 13)],
    }

    # Years for filter dropdown
    years = CallSession.objects.annotate(year=ExtractYear('date')) \
        .values_list('year', flat=True).distinct().order_by('year')

    context = {
        'total_calls': total_calls,
        'prev_total_calls': prev_total_calls,
        'total_delta': abs(delta_total),
        'total_trend': trend_total,
        'common_age': common_age,
        'age_range': age_range,
        'high_risk_percent': high_risk_percent,
        'gender_labels': gender_labels,
        'gender_counts': gender_counts,
        'gender_colors': gender_colors,
        'top_reasons': top_reasons,
        'shift_chart_data': shift_chart_data,
        'selected_year': year,
        'selected_month': selected_month,
        'years': years,
        'month_names': month_names,
    }
    return render(request, 'dashboard.html', context)


@login_required
def export_calls(request):
    year = int(request.GET.get('year', datetime.now().year))
    month = request.GET.get('month')

    calls = CallSession.objects.filter(date__year=year)
    if month:
        calls = calls.filter(date__month=month)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="calls_{year}_{month or "all"}.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Shift", "Caller", "Risk Level", "Age", "Gender", "Status"])

    for c in calls.select_related('caller'):
        writer.writerow([
            c.date,
            c.shift,
            c.caller.name,
            c.risk_level,
            c.caller.age,
            c.caller.get_gender_display(),
            c.caller.get_status_display(),
        ])

    return response


@login_required
def export_soi_excel(request):
    year = int(request.GET.get('year', datetime.now().year))
    month = request.GET.get('month')
    config = SiteConfig.objects.first()

    calls = CallSession.objects.select_related('caller', 'responder').filter(date__year=year)
    if month:
        calls = calls.filter(date__month=month)

    # SHEET 2 – SOI (horizontal table)
    soi_data = []
    for c in calls:
        soi_data.append({
            "Responder": c.responder.first_name if c.responder else "",
            "Tawag Paglaum Code": c.id,
            "Date": c.date.strftime('%Y-%m-%d'),
            "Shift": c.shift,
            "Time Called": c.time_called.strftime('%H:%M') if c.time_called else "",
            "Length of Call": str(c.length_of_call),
            "Time Ended": c.time_ended.strftime('%H:%M') if c.time_ended else "",
            "Name": c.caller.name,
            "Gender": c.caller.get_gender_display(),
            "Status": c.caller.get_status_display(),
            "Age": c.caller.age,
            "Location": c.caller.location,
            "Source of Info": str(c.caller.source_of_info) if c.caller.source_of_info else "",
            "Reason For Calling": c.reasons_for_calling.label if c.reasons_for_calling else "",
            "Intervention": c.interventions,
            "Risk Assessment": c.risk_level,
            "Suicide/ Self Harming Method": c.suicide_methods or "",
            "Other Comments": c.comments or "",
        })
    df_soi = pd.DataFrame(soi_data)

    # Tally data
    def make_summary(df, column, title):
        return df[column].value_counts().reset_index().rename(columns={'index': title, column: 'Count'})

    tally_shift = make_summary(df_soi, "Shift", "Shift")
    tally_gender = make_summary(df_soi, "Gender", "Gender")
    tally_age = make_summary(df_soi, "Age", "Age")
    tally_risk = make_summary(df_soi, "Risk Assessment", "Risk")

    # EXCEL EXPORT
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Sheet 2: SOI
        df_soi.to_excel(writer, sheet_name='SOI', index=False)
        workbook = writer.book

        # Sheet 1: Detailed Report (vertical layout)
        sheet = workbook.create_sheet("Detailed Report", 0)
        writer.sheets["Detailed Report"] = sheet

        # Header
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        sheet.cell(row=1, column=1, value=config.site_name).font = Font(bold=True, size=14)
        if config.logo and os.path.isfile(config.logo.path):
            logo = XLImage(config.logo.path)
            logo.width = 90
            logo.height = 90
            sheet.add_image(logo, "C1")

        row = 3
        for idx, c in enumerate(calls, 1):
            case_fields = [
                ("Case No.", f"TPCEB1-{str(c.id).zfill(3)}"),
                ("Responder No.", c.responder.first_name if c.responder else ""),
                ("Date", c.date.strftime('%Y-%m-%d')),
                ("Time (Duration)", str(c.length_of_call)),
                ("Name", c.caller.name),
                ("Status", c.caller.get_status_display()),
                ("Age", c.caller.age),
                ("Location", c.caller.location),
                ("Source of Info", str(c.caller.source_of_info) if c.caller.source_of_info else ""),
                ("Reason for Calling", c.reasons_for_calling.label if c.reasons_for_calling else ""),
                ("Detailed Report Call", c.ai_summary or "")
            ]
            for label, value in case_fields:
                sheet.cell(row=row, column=1, value=f"{label}:").font = Font(bold=True)
                sheet.cell(row=row, column=2, value=value)
                row += 1
            row += 1

        # Sheet 3: Tally
        sheet = workbook.create_sheet("Tally")
        writer.sheets["Tally"] = sheet
        row = 1
        for name, df in {"Shift": tally_shift, "Gender": tally_gender, "Age": tally_age, "Risk": tally_risk}.items():
            sheet.cell(row=row, column=1, value=name).font = Font(bold=True)
            row += 1
            for col_idx, col in enumerate(df.columns, 1):
                sheet.cell(row=row, column=col_idx, value=col).font = Font(bold=True)
            for tup in df.itertuples(index=False):
                row += 1
                for col_idx, val in enumerate(tup, 1):
                    sheet.cell(row=row, column=col_idx, value=val)

            chart = PieChart() if name == "Gender" else BarChart()
            data = Reference(sheet, min_col=2, min_row=row - len(df) + 1, max_row=row)
            cats = Reference(sheet, min_col=1, min_row=row - len(df) + 1, max_row=row)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            chart.title = f"{name} Distribution"
            if name != "Gender":
                chart.dLbls = DataLabelList()
                chart.dLbls.showVal = True
            if config.primary_color:
                chart.series[0].graphicalProperties.solidFill = config.primary_color.strip("#")
            row += 2
            sheet.add_chart(chart, f"E{row}")
            row += 10

        # Freeze panes & auto-width
        for sname in ["SOI", "Tally"]:
            ws = writer.sheets[sname]
            ws.freeze_panes = "A2"
            for col in ws.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

        # Sheet visibility fix
        for ws in workbook.worksheets:
            ws.sheet_state = "visible"
        workbook.active = workbook.sheetnames.index("SOI")

    buffer.seek(0)
    filename = f"SOI_Report_{calendar.month_name[int(month)] if month else 'All'}_{year}.xlsx"
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def profile_settings(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been updated.')
            return redirect('profile_settings')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'settings/profile_settings.html', {
        'form': form
    })


@staff_member_required
def site_config_settings(request):
    config = SiteConfig.objects.first()
    if request.method == 'POST':
        form = SiteConfigForm(request.POST, request.FILES, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'Site configuration updated successfully.')
            return redirect('site_config_settings')
    else:
        form = SiteConfigForm(instance=config)

    return render(request, 'settings/site_config.html', {
        'form': form
    })


@staff_member_required
def site_config_settings(request):
    config = SiteConfig.objects.first()
    if request.method == 'POST':
        form = SiteConfigForm(request.POST, request.FILES, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'Site configuration updated successfully.')
            return redirect('site_config_settings')
    else:
        form = SiteConfigForm(instance=config)

    return render(request, 'settings/site_config.html', {
        'form': form
    })


@login_required
def settings_dashboard(request):
    user = request.user
    is_admin = user.is_staff
    context = {}

    profile_form = UserProfileForm(instance=user)
    password_form = PasswordChangeForm(user)
    site_form = SiteConfigForm(instance=SiteConfig.objects.first()) if is_admin else None
    today = date.today()
    current_year = today.year
    current_month = today.month

    if request.method == 'POST':
        if 'profile_submit' in request.POST:
            profile_form = UserProfileForm(request.POST, instance=user)
            if profile_form.is_valid():
                profile_form.save()

        elif 'password_submit' in request.POST:
            password_form = PasswordChangeForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)

        elif 'site_submit' in request.POST and is_admin:
            site_form = SiteConfigForm(request.POST, request.FILES, instance=site_form.instance)
            if site_form.is_valid():
                site_form.save()

    # Build a matrix for calendar weeks
    calendar_days = get_calendar_data(current_year, current_month)
    cal = calendar.Calendar(firstweekday=6)  # Sunday as first day
    month_days = cal.monthdatescalendar(current_year, current_month)

    active_tab = request.POST.get("active_tab") or request.GET.get("tab", "profile")

    context = {'profile_form': profile_form, 'password_form': password_form, 'site_form': site_form,
               'calendar_days': calendar_days, 'weekdays': ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat'],
               'month_days': month_days, 'current_month': today.strftime('%B'), 'current_year': current_year,
               'is_admin': is_admin, 'active_tab': active_tab,
               'shift_choices': [(s.id, s.name) for s in ShiftType.objects.all()]}

    return render(request, 'settings.html', context)


@staff_member_required
def delete_holiday(request, pk):
    if request.method == 'POST':
        holiday = get_object_or_404(Holiday, pk=pk)
        holiday.delete()
        messages.success(request, f"{holiday.label} deleted.")
    return redirect(f'{reverse("settings")}?tab=scheduler')


@staff_member_required
def delete_shift_type(request, pk):
    if request.method == 'POST':
        shift = get_object_or_404(ShiftType, pk=pk)
        shift.delete()
        messages.success(request, f"Shift '{shift.name}' deleted.")
    return redirect(f'{reverse("settings")}?tab=scheduler')


@login_required
def shift_calendar(request, year=None, month=None):
    today = datetime.today()
    year = year or today.year
    month = month or today.month
    calendar_days = generate_calendar_matrix(year, month)

    return render(request, 'shift_calendar.html', {
        'calendar_days': calendar_days,
        'month': month,
        'year': year,
    })


SHIFT_ORDER = {
    "Morning": 1,
    "Mid": 2,
    "Afternoon": 3,
    "Night": 4,
    "OFF": 5
}
SHIFT_COLORS = {
    "Morning": "#007bff",  # bg-blue
    "Mid": "#ffc107",  # bg-warning
    "Afternoon": "#fd7e14",  # bg-orange
    "Night": "#343a40",  # bg-dark
    "OFF": "#dc3545"  # bg-danger
}


def calendar_events_api(request):
    user = request.user

    if user.is_staff:
        # Admins see all assignments
        assignments = ShiftAssignment.objects.select_related("shift", "responder").order_by("date")
    else:
        # Responders see only their shifts
        assignments = ShiftAssignment.objects.select_related("shift", "responder").filter(responder=user).order_by(
            "date")

    events = []

    # Holidays
    for h in Holiday.objects.all():
        events.append({
            "title": h.label,
            "start": h.date.isoformat(),
            "color": "#dc3545",  # red
            "allDay": True
        })

    # Sort all assignments by date first, then by SHIFT_ORDER
    sorted_assignments = sorted(
        assignments,
        key=lambda a: (a.date, SHIFT_ORDER.get(a.shift.name.strip().title(), 99))
    )

    # Add sorted assignments to events
    for a in sorted_assignments:
        shift_name = a.shift.name.strip().title()

        # Get first name only
        first_name = a.responder.first_name or a.responder.username

        events.append({
            "title": f"{a.shift.name} ({first_name})",
            "start": a.date.isoformat(),
            "color": SHIFT_COLORS.get(shift_name, "#6c757d"),
            "allDay": True
        })

    return JsonResponse(events, safe=False, encoder=DjangoJSONEncoder)


def get_calendar_data(year, month):
    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])

    # Fetch assignments and holidays for the month
    shift_assignments = ShiftAssignment.objects.filter(date__range=(first_day, last_day)).select_related('shift', 'responder')
    holidays = Holiday.objects.filter(date__range=(first_day, last_day))
    tele_month = MonthlyTeleconsult.objects.filter(month__year=year, month__month=month)

    calendar_days = []
    day_cursor = first_day

    while day_cursor <= last_day:
        assignments_today = shift_assignments.filter(date=day_cursor)
        day_data = {
            'date': day_cursor,
            'is_today': day_cursor == date.today(),
            'is_weekend': day_cursor.weekday() >= 5,
            'holiday': holidays.filter(date=day_cursor).first(),
            'shifts': assignments_today,
            'teleconsult_assigned': tele_month.filter(
                responder__in=[a.responder for a in assignments_today if "tele" in a.shift.name.lower()]
            ).exists()
        }
        calendar_days.append(day_data)
        day_cursor += timedelta(days=1)

    return calendar_days


@login_required
@staff_member_required
def assign_shift(request):
    today = date.today()
    view_mode = request.GET.get("view", "month")
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    # 🗓️ Generate calendar days
    if view_mode == "week" and request.GET.get("start"):
        try:
            start_date = datetime.strptime(request.GET.get("start"), "%Y-%m-%d").date()
        except ValueError:
            start_date = today
        calendar_days = [start_date + timedelta(days=i) for i in range(7)]
    else:
        calendar_days = [date(year, month, d) for d in range(1, monthrange(year, month)[1] + 1)]

    # 📋 Forms
    user = request.user
    is_admin = user.is_staff
    holiday_form = HolidayForm() if is_admin else None
    shift_form = ShiftTypeForm() if is_admin else None
    tele_form = AssignTeleconsultForm() if is_admin else None

    # 📊 Data sources
    holidays = Holiday.objects.order_by("date")
    shift_types = ShiftType.objects.order_by("start_time")
    responders = User.objects.filter(is_superuser=False).order_by("last_name")
    weekend_days = [d for d in calendar_days if d.weekday() >= 5]
    holiday_dates = [h.date for h in holidays if h.date.month == month and h.date.year == year]
    total_off_days = len(set(weekend_days + holiday_dates))
    responder_off_day_counts = defaultdict(int)

    off_shift = ShiftType.objects.filter(name__iexact="Time Off").first()

    if off_shift:
        for assignment in ShiftAssignment.objects.filter(date__year=year, date__month=month, shift=off_shift):
            responder_off_day_counts[assignment.responder_id] += 1

    # 📌 Lookups
    pref_lookup = defaultdict(dict)
    for pref in ShiftPreference.objects.select_related("shift").filter(date__year=year, date__month=month):
        pref_lookup[pref.responder_id][pref.date] = pref

    off_lookup = defaultdict(dict)
    for off in TimeOffRequest.objects.filter(
        start_date__year=year, start_date__month=month,
        status__in=["pending", "approved"]
    ):
        if off.start_date and off.end_date:
            for offset in range((off.end_date - off.start_date).days + 1):
                d = off.start_date + timedelta(days=offset)
                off_lookup[off.responder_id][d] = off
        else:
            off_lookup[off.responder_id][off.start_date] = off

    # ✅ Assigned shifts lookup
    shift_lookup = defaultdict(dict)
    for assignment in ShiftAssignment.objects.filter(date__year=year, date__month=month):
        shift_lookup[assignment.responder_id][assignment.date] = assignment.shift

    # 📍 Auto-assign shifts to teleconsult responder
    try:
        tele_entry = MonthlyTeleconsult.objects.get(month__year=year, month__month=month)
        tele_responder_id = tele_entry.responder_id

        mid_shift = ShiftType.objects.filter(name__iexact="Mid").first()
        off_shift = ShiftType.objects.filter(name__iexact="Time Off").first()

        if tele_responder_id and mid_shift and off_shift:
            auto_assign_teleconsult_preferences(
                responder_id=tele_responder_id,
                calendar_days=calendar_days,
                mid_shift=mid_shift,
                off_shift=off_shift,
            )
    except MonthlyTeleconsult.DoesNotExist:
        tele_entry = None

    # 📝 Handle form submissions
    if request.method == "POST":
        if "holiday_submit" in request.POST and is_admin:
            holiday_form = HolidayForm(request.POST)
            if holiday_form.is_valid():
                holiday_form.save()
                messages.success(request, "Holiday added.")
                return redirect(f"{reverse('settings')}?tab=scheduler")

        elif "shift_submit" in request.POST and is_admin:
            shift_form = ShiftTypeForm(request.POST)
            if shift_form.is_valid():
                shift_form.save()
                messages.success(request, "Shift type added.")
                return redirect(f"{reverse('settings')}?tab=scheduler")

        elif "teleconsult_submit" in request.POST and is_admin:
            tele_form = AssignTeleconsultForm(request.POST)
            if tele_form.is_valid():
                tele_form.save()
                messages.success(request, "Teleconsultation responder assigned.")
                return redirect(f"{reverse('settings')}?tab=scheduler")
            else:
                print("Teleconsult form errors:", tele_form.errors)

    # 📆 Month & year options
    from calendar import month_name
    month_options = [(i, month_name[i]) for i in range(1, 13)]
    year_range = range(today.year - 2, today.year + 3)

    shift_pref_counts = defaultdict(lambda: defaultdict(int))

    prefs = ShiftPreference.objects.filter(date__year=year, date__month=month)

    for pref in prefs:
        shift_name = pref.shift.name
        shift_pref_counts[pref.date][shift_name] += 1

    context = {
        "calendar_days": calendar_days,
        "months": month_options,
        "year_range": year_range,
        "responders": responders,
        "shift_css_classes": {
            "Morning": "bg-blue text-light",
            "Mid": "bg-warning text-dark",
            "Afternoon": "bg-orange text-light",
            "Night": "bg-dark text-light",
            "OFF": "bg-danger text-light",
        },
        "current_month": date(year, month, 1).strftime('%B'),
        "current_month_num": month,
        "current_year": year,
        "view_mode": view_mode,
        "holiday_form": holiday_form,
        "holidays": holidays,
        "weekend_days": weekend_days,
        "total_off_days": total_off_days,
        "responder_off_day_counts": responder_off_day_counts,
        "shift_form": shift_form,
        "tele_form": tele_form,
        "shift_types": shift_types,
        "pref_lookup": pref_lookup,
        "off_lookup": off_lookup,
        "tele_lookup": {tele_responder_id: True} if tele_responder_id else {},
        "shift_lookup": shift_lookup,
        "mid_shift": mid_shift,
        "off_shift": off_shift,
        "shift_pref_counts": shift_pref_counts,
    }

    return render(request, "scheduler/assign_shift.html", context)


def auto_assign_teleconsult_preferences(responder_id, calendar_days, mid_shift, off_shift):
    for day in calendar_days:
        already_pref = ShiftPreference.objects.filter(responder_id=responder_id, date=day).exists()
        if not already_pref:
            shift = off_shift if day.weekday() >= 5 else mid_shift
            ShiftPreference.objects.create(
                responder_id=responder_id,
                date=day,
                shift=shift
            )


@login_required
def request_shift(request):
    if request.method == 'POST':
        user = request.user
        shift_id = request.POST.get('shift')
        timeoff_days = request.POST.getlist('timeoff_days')
        reason = request.POST.get('reason', '')
        month_str = request.POST.get('month')  # Format: YYYY-MM
        leave_start = request.POST.get('leave_start')
        leave_end = request.POST.get('leave_end')

        try:
            # Extract month and year from input
            target_date = datetime.strptime(month_str, "%Y-%m")
            year = target_date.year
            month = target_date.month
            shift_type = ShiftType.objects.get(id=shift_id)
        except (ValueError, ShiftType.DoesNotExist):
            messages.error(request, "Invalid shift or month format.")
            return redirect('settings')

        # Save shift preference for each day of the month
        num_days = calendar.monthrange(year, month)[1]
        for day in range(1, num_days + 1):
            date_obj = date(year, month, day)
            ShiftPreference.objects.update_or_create(
                responder=user,
                date=date_obj,
                shift=shift_type
            )

        # Handle weekday-based time-offs (e.g., Sat & Sun)
        weekdays_index = [list(calendar.day_abbr).index(day.title()) for day in timeoff_days]
        timeoff_requests = []
        for day in range(1, num_days + 1):
            d = date(year, month, day)
            if d.weekday() in weekdays_index:
                timeoff_requests.append(TimeOffRequest(
                    responder=user,
                    start_date=d,
                    end_date=d,
                    shift=None,  # full-day off
                    reason=reason,
                    status='pending'
                ))
        TimeOffRequest.objects.bulk_create(timeoff_requests)

        # Handle optional leave request (start to end)
        if leave_start and leave_end:
            try:
                start = datetime.strptime(leave_start, '%Y-%m-%d').date()
                end = datetime.strptime(leave_end, '%Y-%m-%d').date()
                for d in daterange(start, end):
                    TimeOffRequest.objects.update_or_create(
                        responder=user,
                        start_date=start,
                        end_date=end,
                        shift=None,
                        defaults={'reason': reason, 'status': 'pending'}
                    )
            except ValueError:
                messages.warning(request, "Invalid leave date range.")

        messages.success(request, "Your request has been submitted.")
        return redirect('settings')

    return redirect('settings')


@require_POST
def submit_shifts(request):
    print("submit_shifts view was called")
    user = request.user  # Admin assigning shifts

    for key, value in request.POST.items():
        if key.startswith("assignments-") and value:
            try:
                _, responder_id, date_str = key.split("-", 2)
                date = datetime.strptime(date_str, "%Y-%m-%d").date()
                shift_id = int(value)

                print(f"Assigning: Responder {responder_id} → Shift {shift_id} on {date}, assigned by {user}")

                ShiftAssignment.objects.update_or_create(
                    responder_id=responder_id,
                    date=date,
                    defaults={
                        "shift_id": shift_id,
                        "assigned_by": user
                    }
                )
            except (ValueError, TypeError) as e:
                print(f"Skipping malformed input: {key}={value} → {e}")
                continue

    messages.success(request, "Shift assignments updated.")
    return redirect(reverse("settings") + "?tab=scheduler")




